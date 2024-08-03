import os
import json
import matplotlib.pyplot as plt
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from BinPacking import Rectangle
import networkx as nx
import numpy as np

def remove_txt_extension(filename):
    if filename.endswith(".ins2D"):
        return filename[:-6]
    return filename

def read_items(file_path):
    file_path += ".ins2D"
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            if not content:
                raise ValueError("File is empty")
            lines = content.strip().split('\n')
            item_count = int(lines[0].strip())
            bin_width, bin_height = map(int, lines[1].strip().split())

            items = []
            id = -1
            for line in lines[2:2 + item_count]:
                id += 1
                item_width, item_height = map(int, line.strip().split())
                items.append(Rectangle(id, item_width, item_height))  # Create Rectangle objects
            return bin_width, bin_height, items
    except FileNotFoundError:
        raise FileNotFoundError(f"Unable to open file: {file_path}")

def visualize_bins(bins, folder_path, index):
    if not bins:
        print("No bins to visualize.")
        return

    bins_per_row = 5
    num_bins = len(bins)
    num_rows = (num_bins + bins_per_row - 1) // bins_per_row
    plt.figure(figsize=(15, num_rows * 3))

    colors = ['c', 'orange', 'purple', 'g', 'm', 'r', 'y', 'b', 'brown', 'k']
    for i, bin in enumerate(bins):
        plt.subplot(num_rows, bins_per_row, i + 1)
        for j, item in enumerate(bin.content):
            rect = plt.Rectangle((item.x, item.y), item.width, item.height, fc=colors[j % len(colors)], ec='k')
            plt.gca().add_patch(rect)
            # plt.text(item.x + item.width / 2, item.y + item.height / 2, f'id: {item.id}', ha='center', va='center')
        # print(f"w={bin.height}, h={bin.width}")
        # Highlight unoccupied areas
        for space in bin.unoccupied_spaces:
            ux1, ux2, uy1, uy2 = space.to_tuple()  
            plt.gca().add_patch(plt.Rectangle((ux1, uy1), ux2 - ux1, uy2 - uy1, facecolor='none', ec='red', hatch='/'))
    
    
        plt.xlim(0, bin.width)
        plt.ylim(0, bin.height)
        plt.title(f'Bin {i + 1}')
        plt.xticks([])
        plt.yticks([])

    plt.tight_layout()
    create_directory_if_not_exists(folder_path)
    plt.savefig(os.path.join(folder_path, f"visualization-{index}.png"))
    plt.close()
    


def plot_objectives(best_objectives, avg_objectives, file_path):
    plt.figure()
    plt.plot(best_objectives, label='Best Objectives')
    plt.plot(avg_objectives, label='Average Objectives')
    plt.title("Objectives Over Iterations")
    plt.xlabel("Generation")
    plt.ylabel("Objective Value")
    plt.legend()
    plt.savefig(os.path.join(file_path, "objectives.png"))
    # plt.show()
    plt.close()


def plot_ants(number_ants, file_path):
    plt.figure()
    plt.plot(number_ants, label='Number of Ants')
    plt.title("Ants Over Iterations")
    plt.xlabel("Iteration")
    plt.ylabel("number of ants")
    plt.legend()
    plt.savefig(os.path.join(file_path, "ants.png"))
    # plt.show()
    plt.close()


def plot_acs(ACS_probs, file_path):
    plt.figure()
    plt.plot(ACS_probs, label='ACS prob')
    plt.title("ACS Prob Over Iterations")
    plt.xlabel("Iteration")
    plt.ylabel("ACS prob")
    plt.legend()
    plt.savefig(os.path.join(file_path, "probs.png"))
    # plt.show()
    plt.close()

def create_directory_if_not_exists(folder_path):
    Path(folder_path).mkdir(parents=True, exist_ok=True)


def read_config(folder_path):
    with open(folder_path, 'r') as file:
        return json.load(file)


def read_excel():
    wb = load_workbook('solutions.xlsx')
    sheet = wb.active
    dataset_names = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        dataset_names.append(remove_txt_extension(row[0]))
    return dataset_names


def read_excel_best_solutions():
    wb = load_workbook('solutions.xlsx')
    sheet = wb.active
    best_values = []
    # Adjust the min_col and max_col to read the third column
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        # Since we are reading only one column, row[0] will give us the value
        best_values.append(row[0])
    return best_values

def write_excel_bins(bins_size):
    wb = load_workbook("solutions_new.xlsx")
    sheet = wb.active
    column = sheet.max_column + 1
    for i, size in enumerate(bins_size, start=1):
        sheet.cell(row=i, column=column, value=size)
    wb.save("solutions_new.xlsx")


def write_excel_time( execution_times):
    wb = load_workbook("solutions_time.xlsx")
    sheet = wb.active
    column = sheet.max_column + 1
    for i, time in enumerate(execution_times, start=1):
        sheet.cell(row=i, column=column, value=time)
    wb.save("solutions_time.xlsx")



def plot_pheromone_graph(pheromones, plot_path, show_plot=False):
    matrix = np.array(pheromones)
    G = nx.from_numpy_array(matrix)
    fig, ax = plt.subplots()
    pos = nx.spring_layout(G, weight='weight')
    nx.draw_networkx_nodes(G, pos, node_size=500, node_color='skyblue', ax=ax)

    edges = G.edges(data=True)
    edge_colors = [data['weight'] for _, _, data in edges if data['weight'] > 0]
    
    if edge_colors:
        cmap = plt.cm.Blues
        vmin = min(edge_colors)
        vmax = max(edge_colors)
        edges = nx.draw_networkx_edges(G, pos, edge_cmap=cmap, edge_color=edge_colors, width=2, ax=ax)
        nx.draw_networkx_labels(G, pos, font_size=10, font_family="sans-serif", ax=ax)

        # if vmin != vmax:
        #     sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(vmin=vmin, vmax=vmax))
        #     sm.set_array([])
        #     plt.colorbar(sm, ax=ax, orientation='vertical')
    else:
        print("No edges with non-zero weights to display.")

    plt.axis('off')
    plt.savefig(plot_path + "/graph.png")
    if show_plot:
        plt.show()
    plt.close(fig)  # Close the figure to free up memory
    return pos
