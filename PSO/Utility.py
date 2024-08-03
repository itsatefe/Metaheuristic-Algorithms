import os
import json
import matplotlib.pyplot as plt
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from BinPacking import Rectangle
import networkx as nx
import numpy as np
import itertools
import seaborn as sns
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import squareform

def remove_txt_extension(filename):
    if filename.endswith(".ins2D"):
        return filename[:-6]
    return filename

def read_items(file_path):
    file_path += ".ins2D"
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            if not content:
                raise ValueError("File is empty")
            lines = content.strip().split('\n')
            item_count = int(lines[0].strip())
            bin_width, bin_height = map(int, lines[1].strip().split())

            items = []
            id = -1
            for line in lines[2:2 + item_count]:
                id += 1
                item_width, item_height = map(int, line.strip().split())
                items.append(Rectangle(id, item_width, item_height))  # Create Rectangle objects
            return bin_width, bin_height, items
    except FileNotFoundError:
        raise FileNotFoundError(f"Unable to open file: {file_path}")

def visualize_bins(bins, folder_path, index):
    if not bins:
        print("No bins to visualize.")
        return

    bins_per_row = 5
    num_bins = len(bins)
    num_rows = (num_bins + bins_per_row - 1) // bins_per_row
    plt.figure(figsize=(15, num_rows * 3))

    colors = ['c', 'orange', 'purple', 'g', 'm', 'r', 'y', 'b', 'brown', 'k']
    for i, bin in enumerate(bins):
        plt.subplot(num_rows, bins_per_row, i + 1)
        for j, item in enumerate(list(bin.content.values())):
            rect = plt.Rectangle((item.x, item.y), item.width, item.height, fc=colors[j % len(colors)], ec='k')
            plt.gca().add_patch(rect)
            # plt.text(item.x + item.width / 2, item.y + item.height / 2, f'id: {item.id}', ha='center', va='center')
        # print(f"w={bin.height}, h={bin.width}")
        # Highlight unoccupied areas
        for space in bin.unoccupied_spaces:
            ux1, ux2, uy1, uy2 = space.to_tuple()  
            plt.gca().add_patch(plt.Rectangle((ux1, uy1), ux2 - ux1, uy2 - uy1, facecolor='none', ec='red', hatch='/'))
    
    
        plt.xlim(0, bin.width)
        plt.ylim(0, bin.height)
        plt.title(f'Bin {i + 1}')
        plt.xticks([])
        plt.yticks([])

    plt.tight_layout()
    create_directory_if_not_exists(folder_path)
    plt.savefig(os.path.join(folder_path, f"visualization-{index}.png"))
    plt.close()


def plot_detailed_objectives(detailed_objectives, folder_path):
    num_iterations = len(detailed_objectives)
    num_islands = len(detailed_objectives[0])
    
    # Calculate rows needed for two columns
    num_rows = (num_islands + 1) // 2 if num_islands % 2 == 1 else num_islands // 2
    
    # Setup the figure and subplots with two columns where applicable
    fig, axes = plt.subplots(nrows=num_rows, ncols=2, figsize=(16, num_rows * 3), sharex=True)
    
    # Ensure axes is always a 2D array for consistency in indexing
    axes = np.array(axes).reshape(num_rows, -1)
    
    # Iterate over each island to plot its data
    for i in range(num_islands):
        # Calculate row and column index for the current subplot
        row_idx = i // 2
        col_idx = i % 2

        # Extract avg and best objectives for each iteration for the current island
        avg_objectives = list(itertools.chain.from_iterable(
            detailed_objectives[iteration][i][0] for iteration in range(num_iterations)))
        best_objectives = list(itertools.chain.from_iterable(
            detailed_objectives[iteration][i][1] for iteration in range(num_iterations)))

        # Plotting each objective on its respective subplot
        ax = axes[row_idx, col_idx]
        ax.plot(best_objectives, label='Best Objectives')
        ax.plot(avg_objectives, label='Average Objectives')
        ax.set_title(f'Island {i + 1}')
        ax.set_ylabel('Objective Value')
        ax.grid(True)
        ax.legend()

    # Turn off unused subplot if number of islands is odd
    if num_islands % 2 == 1:
        axes[-1, -1].axis('off')  # Turn off the last subplot in the last row

    # Set common labels and adjust layout to prevent overlap
    plt.xlabel('Iteration')
    plt.tight_layout()
    
    # Save the figure
    plt.savefig(os.path.join(folder_path, "detailed_objectives.png"))
    plt.close()

def plot_objectives(best_objectives, avg_objectives, file_path):
    plt.figure()
    plt.plot(best_objectives, label='Best Objectives')
    plt.plot(avg_objectives, label='Average Objectives')
    plt.title("Objectives Over Iterations")
    plt.xlabel("Generation")
    plt.ylabel("Objective Value")
    plt.legend()
    plt.savefig(os.path.join(file_path, "objectives.png"))
    # plt.show()
    plt.close()

def create_directory_if_not_exists(folder_path):
    Path(folder_path).mkdir(parents=True, exist_ok=True)


def read_config(folder_path):
    with open(folder_path, 'r') as file:
        return json.load(file)


def read_excel():
    wb = load_workbook('solutions.xlsx')
    sheet = wb.active
    dataset_names = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        dataset_names.append(remove_txt_extension(row[0]))
    return dataset_names


def write_excel_bins(bins_size):
    wb = load_workbook("solutions_new.xlsx")
    sheet = wb.active
    column = sheet.max_column + 1
    for i, size in enumerate(bins_size, start=1):
        sheet.cell(row=i, column=column, value=size)
    wb.save("solutions_new.xlsx")

def read_excel_best_solutions():
    wb = load_workbook('solutions.xlsx')
    sheet = wb.active
    best_values = []
    # Adjust the min_col and max_col to read the third column
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        # Since we are reading only one column, row[0] will give us the value
        best_values.append(row[0])
    return best_values

def write_excel_time( execution_times):
    wb = load_workbook("solutions_time.xlsx")
    sheet = wb.active
    column = sheet.max_column + 1
    for i, time in enumerate(execution_times, start=1):
        sheet.cell(row=i, column=column, value=time)
    wb.save("solutions_time.xlsx")


def diversity_heatmap(diversity, folder_path):
    diversity_matrix = np.array(diversity)
    # Create a heatmap
    sns.heatmap(diversity_matrix, annot=True, cmap='coolwarm', fmt=".2f", cbar=False)
    plt.title('Swarm Diversity Heatmap')
    plt.xlabel('Particles')
    plt.ylabel('Particles')
    plt.savefig(os.path.join(folder_path, "diversity_matrix.png"))
    plt.close()

    # Check if the matrix is symmetric and use only the upper triangle minus the diagonal if true
    if np.allclose(diversity_matrix, diversity_matrix.T):
        # Convert the full matrix to a condensed matrix format
        condensed_matrix = squareform(diversity_matrix, checks=False)
    else:
        # If the matrix is not symmetric, handle the situation (or throw an error)
        raise ValueError("The diversity matrix is not symmetric.")

    # Perform hierarchical clustering
    Z = linkage(condensed_matrix, 'ward')
    # Plot dendrogram
    dendrogram(Z, labels=[f"{i+1}" for i in range(len(diversity_matrix))])
    plt.title('Dendrogram of Clustering Results')
    plt.xlabel('Clustering Result')
    plt.ylabel('Distance')
    plt.savefig(os.path.join(folder_path, "diversity_dendrogram.png"))
    plt.close()

